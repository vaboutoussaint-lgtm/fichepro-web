"""
FichePro Web — Backend Flask
Multi-tenant SaaS : chaque client s'authentifie avec sa clé de licence.
Stockage : Supabase (PostgreSQL + Storage).
"""
import os, io, json, hashlib, uuid
from datetime import datetime, date
from functools import wraps

from flask import (
    Flask, request, jsonify, render_template,
    send_file, session, redirect, url_for
)
from flask_cors import CORS
from supabase import create_client, Client
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ─── Configuration ────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "fichepro-secret-2025-change-me")
CORS(app, supports_credentials=True)

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY) if SUPABASE_URL else None

SECTIONS = ["A", "B", "C", "D", "E", "F", "G", "H"]

# ─── Clé admin illimitée ──────────────────────────────────────────────────────
ADMIN_KEY = "WIL03943487HAY"   # votre clé admin personnelle — toujours valide

# ─── Réplication de l'algorithme VBA CreateSimpleHash ─────────────────────────
def _vba_hash(txt: str) -> str:
    """Reproduit CreateSimpleHash du VBA Excel"""
    h = 0
    for ch in txt:
        h = (h * 31 + ord(ch)) % 100000
    return str(h).zfill(5)

def decode_vba_key(key: str):
    """
    Décode une clé générée par le VBA.
    Format :  WIL + hash(5) + duree(3) + tag(3) + HAY  → 17 chars
          ou  WIL + hash(5) + duree(3) + HAY            → 14 chars (ancien)
    Retourne (duration_days: int, hash_ok: bool) ou None si format invalide.
    Le hash ne peut pas être vérifié sans la date d'activation (non stockée
    dans la clé), donc on vérifie uniquement le format et la durée.
    """
    if not (key.startswith("WIL") and key.endswith("HAY")):
        return None
    inner = key[3:-3]          # retire WIL et HAY
    if len(inner) == 8:        # hash5 + dur3
        dur_str = inner[5:8]
    elif len(inner) == 11:     # hash5 + dur3 + tag3
        dur_str = inner[5:8]
    else:
        return None
    try:
        return int(dur_str)    # nombre de jours
    except ValueError:
        return None

# ─── Auth helpers ─────────────────────────────────────────────────────────────
def get_license():
    """Récupère la clé de licence depuis la session ou le header."""
    return session.get("license_key") or request.headers.get("X-License-Key")

def check_license_validity(record):
    """
    Vérifie la validité d'une licence.
    Modèle : le décompte commence à la première activation (activated_at).
    - Si activated_at est null → première connexion → on active maintenant.
    - Expiry = activated_at + duration_days jours.
    - Si duration_days est null → licence illimitée.
    Retourne (valide: bool, message: str, expiry_date: date|None)
    """
    today = date.today()
    activated_at = record.get("activated_at")
    duration_days = record.get("duration_days")  # peut être None = illimité

    if not activated_at:
        # Première activation — on enregistre aujourd'hui
        supabase.table("licenses").update({"activated_at": today.isoformat()}).eq("license_key", record["license_key"]).execute()
        activated_at = today.isoformat()

    if duration_days:
        act_date = datetime.strptime(activated_at[:10], "%Y-%m-%d").date()
        expiry = act_date + __import__('datetime').timedelta(days=int(duration_days))
        jours_restants = (expiry - today).days
        if today > expiry:
            return False, f"Licence expirée depuis le {expiry.strftime('%d/%m/%Y')}", expiry
        return True, f"{jours_restants} jour(s) restant(s)", expiry
    else:
        return True, "Licence illimitée", None

def require_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        key = get_license()
        if not key:
            return jsonify({"error": "Non authentifié"}), 401

        # Clé admin → toujours valide, pas besoin de Supabase
        if key == ADMIN_KEY:
            request.license = {"license_key": key, "client_name": "WILFRIED - Admin",
                               "duration_days": None, "activated_at": None}
            return f(*args, **kwargs)

        lic = supabase.table("licenses").select("*").eq("license_key", key).execute()
        if not lic.data:
            # Clé non enregistrée → refusée (sécurité : pas d'auto-enregistrement)
            return jsonify({"error": "Clé de licence non reconnue. Contactez votre revendeur."}), 403

        record = lic.data[0]
        valide, message, _ = check_license_validity(record)
        if not valide:
            return jsonify({"error": message}), 403
        request.license = record
        return f(*args, **kwargs)
    return decorated

# ─── Auth routes ──────────────────────────────────────────────────────────────
@app.route("/api/login", methods=["POST"])
def login():
    try:
        data = request.get_json() or {}
        key = (data.get("license_key") or "").strip().upper()
        if not key:
            return jsonify({"ok": False, "error": "Clé manquante"}), 400

        if not supabase:
            return jsonify({"ok": False, "error": "Supabase non configuré (SUPABASE_URL manquant)"}), 500

        # ── Clé admin illimitée ──────────────────────────────────────────────
        if key == ADMIN_KEY:
            record = {"license_key": key, "client_name": "WILFRIED - Admin",
                      "duration_days": None, "activated_at": None}
            message, expiry = "Licence Admin illimitée", None
        else:
            # ── Chercher dans Supabase ───────────────────────────────────────
            lic = supabase.table("licenses").select("*").eq("license_key", key).execute()

            if lic.data:
                record = lic.data[0]
                valide, message, expiry = check_license_validity(record)
                if not valide:
                    return jsonify({"ok": False, "error": message}), 403
            else:
                # Clé non enregistrée → refusée
                return jsonify({"ok": False, "error": "Clé de licence non reconnue. Contactez votre revendeur."}), 403
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": f"Erreur serveur: {str(e)}", "trace": traceback.format_exc()}), 500

    session["license_key"] = key
    try:
        prods = supabase.table("producers").select("id").eq("license_key", key).execute()
        has_producers = len(prods.data) > 0
    except Exception:
        has_producers = False

    return jsonify({
        "ok": True,
        "client_name": record.get("client_name", "Client"),
        "expiry_date": expiry.isoformat() if expiry else None,
        "licence_info": message,
        "has_producers": has_producers,
        "license_key": key
    })

@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"ok": True})

# ─── Route admin : enregistrer une clé VBA générée ───────────────────────────
@app.route("/api/admin/register-key", methods=["POST"])
def admin_register_key():
    """Seul l'admin (WIL03943487HAY) peut enregistrer une nouvelle clé."""
    caller = session.get("license_key") or request.headers.get("X-License-Key", "")
    if caller != ADMIN_KEY:
        return jsonify({"ok": False, "error": "Accès refusé"}), 403

    data = request.get_json() or {}
    key          = (data.get("license_key") or "").strip().upper()
    client_name  = (data.get("client_name") or "").strip()
    activation   = (data.get("activated_at") or "").strip()   # YYYY-MM-DD
    duration_days= data.get("duration_days")

    if not key or not key.startswith("WIL") or not key.endswith("HAY"):
        return jsonify({"ok": False, "error": "Format de clé invalide (doit commencer par WIL et finir par HAY)"}), 400

    # Vérifier que la durée est cohérente avec la clé VBA
    dur_from_key = decode_vba_key(key)
    if duration_days is None:
        duration_days = dur_from_key
    if duration_days is None:
        return jsonify({"ok": False, "error": "Durée introuvable dans la clé"}), 400

    if not activation:
        activation = date.today().isoformat()

    # Vérifier si déjà existante
    existing = supabase.table("licenses").select("license_key").eq("license_key", key).execute()
    if existing.data:
        return jsonify({"ok": False, "error": "Cette clé est déjà enregistrée"}), 409

    rec = {
        "license_key"  : key,
        "client_name"  : client_name or f"Client ({key[:6]}...)",
        "duration_days": int(duration_days),
        "activated_at" : activation,
    }
    supabase.table("licenses").insert(rec).execute()
    from datetime import timedelta
    expiry = datetime.strptime(activation, "%Y-%m-%d").date() + timedelta(days=int(duration_days))
    return jsonify({"ok": True, "key": key, "client": rec["client_name"],
                    "expiry": expiry.isoformat()})

@app.route("/api/me")
@require_auth
def me():
    key = get_license()
    prods = supabase.table("producers").select("id").eq("license_key", key).execute()
    fiches = supabase.table("fiches").select("id").eq("license_key", key).execute()
    return jsonify({
        "client_name": request.license.get("client_name"),
        "expiry_date": request.license.get("expiry_date"),
        "has_producers": (len(prods.data)) > 0,
        "producers_count": len(prods.data),
        "fiches_count": len(fiches.data)
    })

# ─── Upload Excel ──────────────────────────────────────────────────────────────
@app.route("/api/upload-excel", methods=["POST"])
@require_auth
def upload_excel():
    key = get_license()
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "Aucun fichier"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsm", ".xlsx", ".xls")):
        return jsonify({"ok": False, "error": "Format non supporté (xlsm/xlsx attendu)"}), 400

    try:
        wb = openpyxl.load_workbook(f, read_only=True, keep_vba=False)
        producers_data = []
        # Chercher l'onglet Producteurs
        target_sheet = None
        for name in wb.sheetnames:
            if any(kw in name.lower() for kw in ["base", "prod", "registre", "agricult", "parc", "ra"]):
                target_sheet = wb[name]
                break
        if not target_sheet and wb.sheetnames:
            target_sheet = wb.active

        header_row = None
        col_map = {}
        is_base_parcelles = target_sheet and "base" in target_sheet.title.lower()

        if target_sheet:
            if is_base_parcelles:
                # Structure fixe "Base Parcelles" — index 0-based (COL_ constants du VBA/engine.py)
                # CODE=0, PARC=1, NOM=2, SECT=3, ESTIM=8
                col_map = {
                    "code": 0, "nom": 2, "section": 3,
                    "estim": 8,
                    "surface": None, "telephone": None,
                }
                header_row = True
            else:
                for row in target_sheet.iter_rows(min_row=1, max_row=5, values_only=True):
                    if row and any(c for c in row if c):
                        headers = [str(c).lower().strip() if c else "" for c in row]
                        for i, h in enumerate(headers):
                            if any(k in h for k in ["code", "num", "id"]):
                                col_map["code"] = i
                            elif any(k in h for k in ["nom", "name", "prenom"]):
                                col_map.setdefault("nom", i)
                            elif "section" in h or "sect" in h:
                                col_map["section"] = i
                            elif any(k in h for k in ["surface", "superf", "ha", "hect"]):
                                col_map["surface"] = i
                            elif any(k in h for k in ["tel", "phone", "mobile"]):
                                col_map["telephone"] = i
                        if col_map:
                            header_row = True
                            break

            for row in target_sheet.iter_rows(min_row=2 if header_row else 1, values_only=True):
                if not row or not any(row):
                    continue
                def val(k):
                    idx = col_map.get(k)
                    return str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] else ""
                code = val("code")
                nom = val("nom")
                if not code and not nom:
                    continue
                section_raw = val("section")
                section = section_raw.lstrip("§").strip() or "A"
                def _f(k):
                    v = val(k)
                    try: return float(v) if v else 0.0
                    except: return 0.0
                producers_data.append({
                    "license_key": key,
                    "code":        code or f"P{len(producers_data)+1:04d}",
                    "nom":         nom or "Inconnu",
                    "section":     section,
                    "surface_ha":  _f("surface"),
                    "telephone":   val("telephone"),
                    "estim":       _f("estim"),
                })

        if not producers_data:
            return jsonify({"ok": False, "error": "Aucun producteur trouvé dans le fichier. Essayez l'import RA."}), 400

        # Supprimer anciens producteurs du client puis insérer
        supabase.table("producers").delete().eq("license_key", key).execute()
        batch_size = 50
        for i in range(0, len(producers_data), batch_size):
            supabase.table("producers").insert(producers_data[i:i+batch_size]).execute()

        return jsonify({"ok": True, "imported": len(producers_data), "message": f"{len(producers_data)} producteurs importés"})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# ─── Import RA (CSV/Excel simple) ─────────────────────────────────────────────
@app.route("/api/import-ra", methods=["POST"])
@require_auth
def import_ra():
    key = get_license()
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "Aucun fichier"}), 400
    f = request.files["file"]
    fname = f.filename.lower()

    try:
        producers_data = []
        if fname.endswith(".csv"):
            import csv
            content = f.read().decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(content), delimiter=";")
            for row in reader:
                def cv(keys):
                    for k in keys:
                        for rk, rv in row.items():
                            if k in rk.lower() and rv:
                                return str(rv).strip()
                    return ""
                code = cv(["code","num","id"])
                nom  = cv(["nom","name","prenom"])
                if not code and not nom:
                    continue
                sec_raw = cv(["section","sect"])
                producers_data.append({
                    "license_key": key,
                    "code": code or f"P{len(producers_data)+1:04d}",
                    "nom": nom or "Inconnu",
                    "section": sec_raw.lstrip("§").strip() or "A",
                    "surface_ha": 0.0,
                    "telephone": cv(["tel","phone","mobile"]),
                })
        else:
            # Déléguer à upload_excel
            return upload_excel()

        if not producers_data:
            return jsonify({"ok": False, "error": "Aucun producteur trouvé"}), 400

        supabase.table("producers").delete().eq("license_key", key).execute()
        for i in range(0, len(producers_data), 50):
            supabase.table("producers").insert(producers_data[i:i+50]).execute()

        return jsonify({"ok": True, "imported": len(producers_data)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# ─── Producteurs ──────────────────────────────────────────────────────────────
@app.route("/api/producers")
@require_auth
def list_producers():
    key = get_license()
    q = request.args.get("q", "")
    section = request.args.get("section", "")
    query = supabase.table("producers").select("*").eq("license_key", key)
    if section:
        query = query.eq("section", section)
    result = query.order("nom").execute()
    data = result.data or []
    if q:
        q_low = q.lower()
        data = [p for p in data if q_low in p.get("nom","").lower() or q_low in p.get("code","").lower()]
    # Enrichir avec nb fiches
    for p in data:
        fc = supabase.table("fiches").select("id").eq("license_key", key).eq("producer_id", p["id"]).execute()
        p["fiches_count"] = len(fc.data)
    return jsonify(data)

@app.route("/api/producers/<prod_id>")
@require_auth
def get_producer(prod_id):
    key = get_license()
    r = supabase.table("producers").select("*").eq("license_key", key).eq("id", prod_id).execute()
    if not r.data:
        return jsonify({"error": "Introuvable"}), 404
    return jsonify(r.data[0])

# ─── Fiches ───────────────────────────────────────────────────────────────────
@app.route("/api/fiches", methods=["GET"])
@require_auth
def list_fiches():
    key = get_license()
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 20))
    producer_id = request.args.get("producer_id")
    section = request.args.get("section")
    month = request.args.get("month")  # YYYY-MM

    query = supabase.table("fiches").select(
        "id, license_key, producer_id, numero_fiche, date_fiche, poids_brut, poids_net, prix_unitaire, montant_total, sections_volumes, calendrier, created_at, producers(nom, code, section)"
    ).eq("license_key", key)

    if producer_id:
        query = query.eq("producer_id", producer_id)
    if month:
        query = query.gte("date_fiche", month + "-01").lte("date_fiche", month + "-31")

    result = query.order("created_at", desc=True).range((page-1)*per_page, page*per_page - 1).execute()
    return jsonify({"data": result.data or [], "page": page})

@app.route("/api/fiches", methods=["POST"])
@require_auth
def create_fiche():
    """
    Crée UNE fiche d'accompagnement groupée.
    Deux cas d'usage :
      1) Fiche manuelle (un seul producteur) : envoyer producer_id + poids_brut
      2) Fiche auto (plusieurs producteurs)  : envoyer producteurs_assignes (liste)
    Pas d'impureté — on travaille en poids brut direct.
    """
    key  = get_license()
    data = request.get_json() or {}

    required = ["date_fiche", "prix_unitaire"]
    for r in required:
        if not data.get(r):
            return jsonify({"ok": False, "error": f"Champ manquant : {r}"}), 400

    prix      = float(data["prix_unitaire"])
    calendrier= data.get("calendrier", "")
    notes     = data.get("notes", "")
    today_dt  = datetime.now()
    count_r   = supabase.table("fiches").select("id").eq("license_key", key).execute()
    num       = f"FP-{today_dt.strftime('%Y%m')}-{len(count_r.data) + 1:04d}"

    # ── Cas 1 : fiche manuelle (un producteur) ───────────────────────────────
    producteurs_assignes = data.get("producteurs_assignes")   # liste ou None
    if not producteurs_assignes:
        pid      = data.get("producer_id")
        poids    = float(data.get("poids_brut", 0))
        if not pid:
            return jsonify({"ok": False, "error": "producer_id ou producteurs_assignes requis"}), 400
        montant  = round(poids * prix, 2)
        prod_r   = supabase.table("producers").select("nom,code,section").eq("license_key", key).eq("id", pid).execute()
        p_info   = prod_r.data[0] if prod_r.data else {}
        producteurs_assignes = [{
            "id": pid, "code": p_info.get("code",""),
            "nom": p_info.get("nom",""), "section": p_info.get("section",""),
            "poids": poids
        }]
    else:
        poids   = sum(float(p.get("poids", 0)) for p in producteurs_assignes)
        montant = round(poids * prix, 2)

    # ── Récap sections ────────────────────────────────────────────────────────
    sections_recap = {}
    for p in producteurs_assignes:
        s = (p.get("section") or "?").lstrip("§").strip()
        sections_recap[s] = round(sections_recap.get(s, 0) + float(p.get("poids", 0)), 2)

    payload_json = json.dumps({
        "type": "multi",
        "producteurs": producteurs_assignes,
        "sections": sections_recap,
        "poids_total": round(poids, 2),
    })

    fiche = {
        "id":              str(uuid.uuid4()),
        "license_key":     key,
        "producer_id":     producteurs_assignes[0].get("id") if producteurs_assignes else None,
        "numero_fiche":    num,
        "date_fiche":      data["date_fiche"],
        "poids_brut":      round(poids, 2),
        "impurete":        0,
        "poids_net":       round(poids, 2),
        "prix_unitaire":   prix,
        "montant_total":   montant,
        "sections_volumes":payload_json,
        "calendrier":      calendrier,
        "notes":           notes,
        "created_at":      datetime.utcnow().isoformat(),
    }
    result = supabase.table("fiches").insert(fiche).execute()
    fiche_out = result.data[0] if result.data else fiche
    fiche_out["producteurs_assignes"] = producteurs_assignes
    fiche_out["sections_recap"]       = sections_recap
    return jsonify({"ok": True, "fiche": fiche_out, "numero": num})

@app.route("/api/fiches/<fiche_id>", methods=["DELETE"])
@require_auth
def delete_fiche(fiche_id):
    key = get_license()
    supabase.table("fiches").delete().eq("license_key", key).eq("id", fiche_id).execute()
    return jsonify({"ok": True})

# ─── Export PDF ───────────────────────────────────────────────────────────────
@app.route("/api/fiches/<fiche_id>/pdf")
@require_auth
def export_pdf(fiche_id):
    key = get_license()
    r = supabase.table("fiches").select(
        "*, producers(nom, code, section, surface_ha, telephone)"
    ).eq("license_key", key).eq("id", fiche_id).execute()
    if not r.data:
        return jsonify({"error": "Fiche introuvable"}), 404

    fiche = r.data[0]
    prod = fiche.get("producers") or {}
    sections_vol = json.loads(fiche.get("sections_volumes") or "{}") if isinstance(fiche.get("sections_volumes"), str) else (fiche.get("sections_volumes") or {})

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=2*cm, rightMargin=2*cm)

    styles = getSampleStyleSheet()
    GREEN  = colors.HexColor("#1a7a3a")
    LGRAY  = colors.HexColor("#f5f5f5")
    DGREEN = colors.HexColor("#145c2b")

    title_style = ParagraphStyle("Title", parent=styles["Normal"],
        fontSize=16, textColor=GREEN, fontName="Helvetica-Bold",
        alignment=TA_CENTER, spaceAfter=4)
    sub_style   = ParagraphStyle("Sub", parent=styles["Normal"],
        fontSize=10, textColor=colors.grey, alignment=TA_CENTER, spaceAfter=12)
    label_style = ParagraphStyle("Label", parent=styles["Normal"],
        fontSize=9, textColor=colors.grey, fontName="Helvetica")
    value_style = ParagraphStyle("Value", parent=styles["Normal"],
        fontSize=11, fontName="Helvetica-Bold")
    sect_style  = ParagraphStyle("Sect", parent=styles["Normal"],
        fontSize=9, fontName="Helvetica", alignment=TA_CENTER)

    story = []

    # En-tête
    story.append(Paragraph("🌿 FICHE D'ACCOMPAGNEMENT", title_style))
    story.append(Paragraph(f"N° {fiche['numero_fiche']}  —  {fiche['date_fiche']}", sub_style))
    story.append(HRFlowable(width="100%", thickness=2, color=GREEN))
    story.append(Spacer(1, 0.4*cm))

    # Infos producteur
    prod_data = [
        ["CODE", prod.get("code","—"), "NOM", prod.get("nom","—")],
        ["SECTION", prod.get("section","—"), "SURFACE", f"{prod.get('surface_ha',0):.2f} ha"],
        ["TÉLÉPHONE", prod.get("telephone","—"), "CALENDRIER", fiche.get("calendrier","—")],
    ]
    pt = Table(prod_data, colWidths=[3*cm, 5*cm, 3*cm, 6*cm])
    pt.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (0,-1), LGRAY),
        ("BACKGROUND", (2,0), (2,-1), LGRAY),
        ("FONTNAME",   (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTNAME",   (2,0), (2,-1), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 9),
        ("TEXTCOLOR",  (0,0), (1,-1), colors.HexColor("#444")),
        ("GRID",       (0,0), (-1,-1), 0.5, colors.lightgrey),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.white, colors.HexColor("#fafafa")]),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
    ]))
    story.append(pt)
    story.append(Spacer(1, 0.5*cm))

    # Poids & Prix
    story.append(Paragraph("▌ PESÉE & TARIFICATION", ParagraphStyle("H2",
        parent=styles["Normal"], fontSize=10, textColor=DGREEN,
        fontName="Helvetica-Bold", spaceAfter=6)))

    poids_data = [
        ["POIDS BRUT", "IMPURETÉ", "POIDS NET", "PRIX UNITAIRE", "MONTANT TOTAL"],
        [
            f"{fiche.get('poids_brut',0):.2f} kg",
            f"{fiche.get('impurete',0):.1f} %",
            f"{fiche.get('poids_net',0):.2f} kg",
            f"{fiche.get('prix_unitaire',0):.0f} F/kg",
            f"{fiche.get('montant_total',0):,.0f} FCFA",
        ]
    ]
    pw = [3.2*cm, 2.8*cm, 3.2*cm, 3.2*cm, 4.6*cm]
    poids_t = Table(poids_data, colWidths=pw)
    poids_t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), GREEN),
        ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,-1), 9),
        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
        ("GRID",          (0,0), (-1,-1), 0.5, colors.lightgrey),
        ("FONTNAME",      (0,1), (-1,1), "Helvetica-Bold"),
        ("BACKGROUND",    (4,1), (4,1), colors.HexColor("#e8f5e9")),
        ("TEXTCOLOR",     (4,1), (4,1), DGREEN),
        ("TOPPADDING",    (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
    ]))
    story.append(poids_t)
    story.append(Spacer(1, 0.5*cm))

    # Volumes par section
    if sections_vol:
        story.append(Paragraph("▌ VOLUMES PAR SECTION", ParagraphStyle("H2",
            parent=styles["Normal"], fontSize=10, textColor=DGREEN,
            fontName="Helvetica-Bold", spaceAfter=6)))
        sv_headers = ["SECTION"] + list(sections_vol.keys()) + ["TOTAL"]
        total_vol = sum(float(v) for v in sections_vol.values())
        sv_row = ["Volumes"] + [f"{v}" for v in sections_vol.values()] + [f"{total_vol:.0f}"]
        sv_t = Table([sv_headers, sv_row])
        sv_t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), GREEN),
            ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,-1), 9),
            ("ALIGN",         (0,0), (-1,-1), "CENTER"),
            ("GRID",          (0,0), (-1,-1), 0.5, colors.lightgrey),
            ("FONTNAME",      (0,1), (-1,1), "Helvetica-Bold"),
            ("TOPPADDING",    (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ]))
        story.append(sv_t)
        story.append(Spacer(1, 0.5*cm))

    # Notes
    if fiche.get("notes"):
        story.append(Paragraph("▌ NOTES", ParagraphStyle("H2",
            parent=styles["Normal"], fontSize=10, textColor=DGREEN,
            fontName="Helvetica-Bold", spaceAfter=4)))
        story.append(Paragraph(fiche["notes"], styles["Normal"]))
        story.append(Spacer(1, 0.4*cm))

    # Signatures
    story.append(Spacer(1, 1*cm))
    sig_data = [["SIGNATURE AGENT", "", "SIGNATURE PRODUCTEUR"]]
    sig_t = Table(sig_data, colWidths=[6*cm, 5*cm, 6*cm])
    sig_t.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("TEXTCOLOR", (0,0), (-1,-1), colors.grey),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("LINEBELOW", (0,0), (0,0), 1, colors.grey),
        ("LINEBELOW", (2,0), (2,0), 1, colors.grey),
        ("TOPPADDING", (0,0), (-1,-1), 20),
    ]))
    story.append(sig_t)

    # Pied de page
    story.append(Spacer(1, 0.8*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey))
    story.append(Paragraph(
        f"Généré par FichePro  —  {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
        ParagraphStyle("Footer", parent=styles["Normal"],
            fontSize=7, textColor=colors.grey, alignment=TA_CENTER, spaceBefore=4)
    ))

    doc.build(story)
    buf.seek(0)
    filename = f"fiche_{fiche['numero_fiche'].replace('/','-')}.pdf"
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=True, download_name=filename)

# ─── Export Excel fiche (tableau simple multi-producteurs) ───────────────────
@app.route("/api/fiches/<fiche_id>/excel")
@require_auth
def export_excel_fiche(fiche_id):
    key = get_license()
    r = supabase.table("fiches").select("*").eq("license_key", key).eq("id", fiche_id).execute()
    if not r.data:
        return jsonify({"error": "Fiche introuvable"}), 404

    fiche = r.data[0]
    # Lire la structure multi-producteurs
    raw = fiche.get("sections_volumes") or "{}"
    if isinstance(raw, str):
        try:
            sv = json.loads(raw)
        except Exception:
            sv = {}
    else:
        sv = raw or {}

    # Compatibilité : fiche multi ou mono
    if sv.get("type") == "multi":
        producteurs = sv.get("producteurs", [])
        sections    = sv.get("sections", {})
    else:
        producteurs = []
        sections    = sv  # ancienne structure {sect: vol}

    GREEN = "1a7a3a"; LGRAY = "f5f5f5"; WHITE = "FFFFFF"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fiche accompagnement"

    def cs(row, col, val, bold=False, bg=None, fg="000000", align="left", size=10, border=True):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, color=fg, size=size)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center")
        if border:
            thin = Side(style="thin", color="CCCCCC")
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return c

    # ── En-tête ───────────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    cs(1, 1, "FICHE D'ACCOMPAGNEMENT", bold=True, bg=GREEN, fg=WHITE,
       align="center", size=13, border=False)
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:F2")
    cs(2, 1, f"N° {fiche['numero_fiche']}   |   Date : {fiche['date_fiche']}   |   Calendrier : {fiche.get('calendrier','')}   |   Prix : {fiche.get('prix_unitaire',0)} F/kg",
       bg="e8f5e9", align="center", size=9, border=False)
    ws.row_dimensions[2].height = 16

    row = 4
    # ── Tableau producteurs ───────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    cs(row, 1, "SÉLECTION PRODUCTEURS — SUIVI LIVRAISONS", bold=True,
       bg="145c2b", fg=WHITE, align="left", size=10, border=False)
    ws.row_dimensions[row].height = 18
    row += 1

    hdrs = ["Code", "Nom Producteur", "Section", "Poids alloué (kg)", "Montant (FCFA)", "Visa / Livraison"]
    prix = float(fiche.get("prix_unitaire", 0))
    for i, h in enumerate(hdrs):
        cs(row, i+1, h, bold=True, bg=GREEN, fg=WHITE, align="center", size=9)
    ws.row_dimensions[row].height = 16
    row += 1

    if producteurs:
        for p in producteurs:
            poids_p = float(p.get("poids", 0))
            cs(row, 1, p.get("code",""), align="center", size=10)
            cs(row, 2, p.get("nom",""), size=10)
            cs(row, 3, p.get("section",""), align="center", size=10)
            cs(row, 4, poids_p, align="center", size=10)
            cs(row, 5, round(poids_p * prix, 0), align="center", size=10)
            cs(row, 6, "", size=10)  # colonne signature vide
            ws.row_dimensions[row].height = 16
            row += 1
    else:
        # Fiche mono-producteur ancienne
        cs(row, 1, "—", align="center"); cs(row, 2, "—"); cs(row, 3, "—")
        cs(row, 4, fiche.get("poids_brut",0), align="center")
        cs(row, 5, fiche.get("montant_total",0), align="center")
        cs(row, 6, ""); row += 1

    # ── Ligne total ───────────────────────────────────────────────────────────
    total_poids = sum(float(p.get("poids",0)) for p in producteurs) if producteurs else float(fiche.get("poids_brut",0))
    total_mont  = round(total_poids * prix, 0)
    ws.merge_cells(f"A{row}:C{row}")
    cs(row, 1, "TOTAL", bold=True, bg=LGRAY, align="right", size=10)
    cs(row, 4, total_poids, bold=True, bg=LGRAY, align="center", size=10)
    cs(row, 5, total_mont,  bold=True, bg=LGRAY, align="center", size=10)
    cs(row, 6, "", bg=LGRAY)
    ws.row_dimensions[row].height = 18
    row += 2

    # ── Récap par section avec point livraison ────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    cs(row, 1, "RÉCAPITULATIF PAR SECTION — POINT DES LIVRAISONS", bold=True,
       bg="145c2b", fg=WHITE, align="left", size=10, border=False)
    ws.row_dimensions[row].height = 18; row += 1

    recap_hdrs = ["Section", "Nb Producteurs", "Volume alloué (kg)", "Montant (FCFA)", "Poids livré (kg)", "Solde restant (kg)"]
    for i, h in enumerate(recap_hdrs):
        cs(row, i+1, h, bold=True, bg=GREEN, fg=WHITE, align="center", size=9)
    ws.row_dimensions[row].height = 16; row += 1

    # Compter nb producteurs et volumes par section
    sect_nb   = {}
    sect_vols = {}
    for p in producteurs:
        s = p.get("section","?")
        sect_nb[s]   = sect_nb.get(s, 0) + 1
        sect_vols[s] = round(sect_vols.get(s, 0) + float(p.get("poids", 0)), 2)

    all_sects = sorted(set(list((sections or {}).keys()) + list(sect_vols.keys())))
    grand_alloue = 0; grand_mont = 0
    for sect in all_sects:
        vol  = sect_vols.get(sect, sections.get(sect, 0))
        nb   = sect_nb.get(sect, 0)
        mont = round(float(vol) * prix, 0)
        grand_alloue += float(vol); grand_mont += mont
        cs(row, 1, sect,  align="center", size=10)
        cs(row, 2, nb,    align="center", size=10)
        cs(row, 3, vol,   align="center", size=10)
        cs(row, 4, mont,  align="center", size=10)
        cs(row, 5, "",    align="center", size=10)   # à remplir à la livraison
        cs(row, 6, "",    align="center", size=10)   # solde calculé manuellement
        ws.row_dimensions[row].height = 15; row += 1

    # Ligne total récap
    ws.merge_cells(f"A{row}:B{row}")
    cs(row, 1, "TOTAL", bold=True, bg=LGRAY, align="right", size=10)
    cs(row, 3, round(grand_alloue, 2), bold=True, bg=LGRAY, align="center", size=10)
    cs(row, 4, round(grand_mont, 0),   bold=True, bg=LGRAY, align="center", size=10)
    cs(row, 5, "", bg=LGRAY)
    cs(row, 6, "", bg=LGRAY)
    ws.row_dimensions[row].height = 18; row += 2

    ws.merge_cells(f"A{row}:F{row}")
    cs(row, 1, f"Généré par FichePro — {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
       align="center", size=8, bg=None, fg="888888", border=False)

    # Largeurs colonnes
    for col_idx, w in enumerate([10, 28, 14, 18, 18, 16], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"fiche_{fiche['numero_fiche'].replace('/','-')}.xlsx"
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)

# ─── Dashboard stats ──────────────────────────────────────────────────────────
@app.route("/api/stats")
@require_auth
def stats():
    key = get_license()
    prods_r = supabase.table("producers").select("*").eq("license_key", key).execute()
    fiches_r = supabase.table("fiches").select("*").eq("license_key", key).execute()
    prods  = prods_r.data or []
    fiches = fiches_r.data or []

    # Nombre de producteurs par section (noms réels, sans §)
    sect_totals = {}
    for p in prods:
        s = (p.get("section") or "?").lstrip("§").strip()
        sect_totals[s] = sect_totals.get(s, 0) + 1

    total_poids = sum(float(f.get("poids_net",0) or 0) for f in fiches)
    total_montant = sum(float(f.get("montant_total",0) or 0) for f in fiches)

    # Top producteurs par fiches
    prod_fiches = {}
    for f in fiches:
        pid = f.get("producer_id","")
        prod_fiches[pid] = prod_fiches.get(pid, 0) + 1
    top_prods = sorted(prod_fiches.items(), key=lambda x: x[1], reverse=True)[:8]
    top_list = []
    prod_map = {p["id"]: p for p in prods}
    for pid, cnt in top_prods:
        p = prod_map.get(pid, {})
        top_list.append({"nom": p.get("nom","?"), "code": p.get("code",""), "fiches": cnt})

    # Fiches du mois
    this_month = datetime.now().strftime("%Y-%m")
    month_fiches = [f for f in fiches if (f.get("date_fiche") or "").startswith(this_month)]

    return jsonify({
        "producers_count": len(prods),
        "fiches_count": len(fiches),
        "fiches_month": len(month_fiches),
        "total_poids_net": round(total_poids, 2),
        "total_montant": round(total_montant, 2),
        "sections": sect_totals,
        "top_producers": top_list,
    })

# ─── Sections distinctes ──────────────────────────────────────────────────────
@app.route("/api/sections")
@require_auth
def get_sections():
    """Retourne la liste des noms de sections distincts pour ce client."""
    key = get_license()
    r = supabase.table("producers").select("section").eq("license_key", key).execute()
    sections = sorted(set(
        p["section"].lstrip("§").strip()
        for p in (r.data or [])
        if p.get("section")
    ))
    return jsonify(sections)

# ─── Génération automatique : UNE fiche multi-producteurs ───────────────────
@app.route("/api/auto-fiches", methods=["POST"])
@require_auth
def auto_fiches():
    """
    Sélectionne intelligemment les producteurs éligibles (scoring VBA simplifié),
    distribue le poids total, puis crée UNE SEULE fiche d'accompagnement
    groupée contenant tous les producteurs sélectionnés.
    Scoring : tauxScore(50%) + volumeScore(30%) + scoreDisp(20%)
    Pas d'impureté — poids brut = poids net.
    """
    key = get_license()
    data = request.get_json() or {}

    poids_total    = float(data.get("poids_total",    0))
    date_fiche     = data.get("date_fiche",     date.today().isoformat())
    prix_unitaire  = float(data.get("prix_unitaire",  0))
    calendrier     = data.get("calendrier",     "")
    notes          = data.get("notes",          "")
    sections_filtre= data.get("sections_filtre",[])
    max_par_prod   = float(data.get("max_par_prod",   1000))

    if poids_total   <= 0: return jsonify({"ok": False, "error": "Poids total invalide"}), 400
    if prix_unitaire <= 0: return jsonify({"ok": False, "error": "Prix unitaire invalide"}), 400

    # ── Producteurs éligibles ─────────────────────────────────────────────────
    prods = supabase.table("producers").select("*").eq("license_key", key).execute().data or []
    if sections_filtre:
        prods = [p for p in prods if (p.get("section") or "").lstrip("§").strip() in sections_filtre]
    if not prods:
        return jsonify({"ok": False, "error": "Aucun producteur éligible dans ces sections"}), 400

    # ── Historique pour scoring ───────────────────────────────────────────────
    fiches_all = (supabase.table("fiches").select("producer_id,created_at")
                  .eq("license_key", key).execute().data or [])
    prod_cnt  = {}
    prod_last = {}
    for f in fiches_all:
        pid = f.get("producer_id","")
        prod_cnt[pid] = prod_cnt.get(pid, 0) + 1
        ts = f.get("created_at","2000-01-01")
        if ts > prod_last.get(pid,""):
            prod_last[pid] = ts

    max_cnt  = max(prod_cnt.values(), default=1) or 1

    # Capacité producteur : ESTIM si dispo, sinon surface_ha * 500 (rendement moyen)
    def capacite(p):
        e = float(p.get("estim") or 0)
        if e > 0:
            return e
        return float(p.get("surface_ha") or 1) * 500

    max_cap = max((capacite(p) for p in prods), default=1) or 1

    def score(p):
        """Réplication du scoring VBA : tauxScore(50%) + volumeScore(30%) + scoreDisp(20%)"""
        pid    = p["id"]
        cnt    = prod_cnt.get(pid, 0)
        cap    = capacite(p)
        last   = prod_last.get(pid, "2000-01-01")
        # tauxScore : favorise les moins utilisés
        taux_s = 1.0 - cnt / max_cnt
        # volumeScore : favorise la plus grande capacité estimée
        vol_s  = cap / max_cap
        # scoreDisp : favorise ceux non sélectionnés depuis longtemps
        try:
            days = (datetime.utcnow() - datetime.fromisoformat(last[:19])).days
        except Exception:
            days = 365
        return 0.5 * taux_s + 0.3 * vol_s + 0.2 * min(days / 365, 1.0)

    prods_sorted = sorted(prods, key=score, reverse=True)

    # ── Distribution du poids — PlafondProducteur (logique VBA) ─────────────
    remaining = poids_total
    producteurs_assignes = []

    for p in prods_sorted:
        if remaining < 0.01:
            break
        cap      = capacite(p)
        # PlafondProducteur = min(capacité, 30%*capacité, max_par_prod, remaining)
        lim30    = cap * 0.30
        plafond  = min(cap, lim30, max_par_prod, remaining)
        if plafond < 1:
            plafond = min(max_par_prod, remaining)
        poids_p = round(min(plafond, remaining), 2)
        if poids_p <= 0:
            continue
        sect = (p.get("section") or "").lstrip("§").strip()
        producteurs_assignes.append({
            "id":      p["id"],
            "code":    p.get("code",""),
            "nom":     p.get("nom",""),
            "section": sect,
            "poids":   poids_p,
        })
        remaining = round(remaining - poids_p, 2)

    if not producteurs_assignes:
        return jsonify({"ok": False, "error": "Impossible de distribuer le poids (vérifiez les plafonds)"}), 400

    # ── Créer UNE seule fiche groupée via create_fiche ────────────────────────
    poids_distribue = round(poids_total - remaining, 2)
    sections_recap  = {}
    for p in producteurs_assignes:
        s = p["section"]
        sections_recap[s] = round(sections_recap.get(s, 0) + p["poids"], 2)

    montant_total = round(poids_distribue * prix_unitaire, 2)
    today_dt = datetime.now()
    count_r  = supabase.table("fiches").select("id").eq("license_key", key).execute()
    num = f"FP-{today_dt.strftime('%Y%m')}-{len(count_r.data) + 1:04d}"

    payload_json = json.dumps({
        "type":        "multi",
        "producteurs": producteurs_assignes,
        "sections":    sections_recap,
        "poids_total": poids_distribue,
    })

    fiche = {
        "id":              str(uuid.uuid4()),
        "license_key":     key,
        "producer_id":     producteurs_assignes[0]["id"],
        "numero_fiche":    num,
        "date_fiche":      date_fiche,
        "poids_brut":      poids_distribue,
        "impurete":        0,
        "poids_net":       poids_distribue,
        "prix_unitaire":   prix_unitaire,
        "montant_total":   montant_total,
        "sections_volumes":payload_json,
        "calendrier":      calendrier,
        "notes":           notes,
        "created_at":      datetime.utcnow().isoformat(),
    }
    supabase.table("fiches").insert(fiche).execute()

    return jsonify({
        "ok":                  True,
        "numero":              num,
        "fiche_id":            fiche["id"],
        "nb_producteurs":      len(producteurs_assignes),
        "poids_distribue":     poids_distribue,
        "poids_restant":       remaining,
        "montant_total":       montant_total,
        "producteurs_assignes":producteurs_assignes,
        "sections_recap":      sections_recap,
    })

# ─── Routes frontend ──────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
