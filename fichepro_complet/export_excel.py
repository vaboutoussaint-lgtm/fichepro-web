"""
FichePro Manager - Export Excel des fiches
Format identique au fichier VBA original
"""
import io
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Couleurs identiques au VBA
C_VERT       = "1B4332"   # Vert foncé entête
C_ORANGE     = "FFC107"   # Orange lignes impaires (proche VBA C_ORANGE=49407)
C_BLANC      = "FFFFFF"
C_JAUNE      = "FFF9C4"   # Ligne 2 (poids total / date)

def generer_export_excel(fiche_data: dict) -> bytes:
    """
    Génère un fichier Excel au format identique au VBA.
    Retourne les bytes du fichier .xlsx
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl non disponible")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Fiche {fiche_data.get('fiche_id', 'export')}"

    # ── Styles ──
    fill_vert   = PatternFill("solid", fgColor=C_VERT)
    fill_orange = PatternFill("solid", fgColor="FF9800")   # lignes impaires
    fill_blanc  = PatternFill("solid", fgColor=C_BLANC)
    fill_jaune  = PatternFill("solid", fgColor="FFF9C4")   # ligne 2

    font_blanc_bold = Font(name="Calibri", size=10, bold=True,  color=C_BLANC)
    font_vert_bold  = Font(name="Calibri", size=10, bold=True,  color=C_VERT)
    font_normal     = Font(name="Calibri", size=10, bold=False, color="1C1C1C")
    font_bold_or    = Font(name="Calibri", size=10, bold=True,  color="7B4F00")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    align_left   = Alignment(horizontal="left",   vertical="center")

    thin = Side(style="thin", color="BDBDBD")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── LIGNE 1 : Entêtes ──
    headers = [
        "Code producteur", "Parcelle", "Producteur", "Section",
        "Poids (kg)", "Nb Sacs", "ID Fiche", "Date Livraison",
        "Prix achat brousse", "Longitude", "Latitude"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = fill_vert
        cell.font      = font_blanc_bold
        cell.alignment = align_center
        cell.border    = border_thin
    ws.row_dimensions[1].height = 22

    # ── LIGNE 2 : Poids total + Date ──
    # Colonne E = label, G = poids total, H = date limite
    ws.cell(row=2, column=5, value="Poids total (kg) :").font = font_bold_or
    ws.cell(row=2, column=5).fill = fill_jaune
    ws.cell(row=2, column=5).alignment = align_right = Alignment(horizontal="right", vertical="center")

    poids_total = fiche_data.get("poids_total", 0)
    ws.cell(row=2, column=7, value=poids_total).font = font_bold_or
    ws.cell(row=2, column=7).fill      = fill_jaune
    ws.cell(row=2, column=7).alignment = align_center
    ws.cell(row=2, column=7).border    = border_thin

    date_ref = fiche_data.get("d_ref", "")
    if date_ref:
        try:
            dv = datetime.strptime(date_ref, "%Y-%m-%d").date() if isinstance(date_ref, str) else date_ref
            ws.cell(row=2, column=8, value=dv)
            ws.cell(row=2, column=8).number_format = "DD/MM/YYYY"
        except:
            ws.cell(row=2, column=8, value=str(date_ref))
    ws.cell(row=2, column=8).font      = font_bold_or
    ws.cell(row=2, column=8).fill      = fill_jaune
    ws.cell(row=2, column=8).alignment = align_center
    ws.cell(row=2, column=8).border    = border_thin

    for c in range(1, 12):
        cell = ws.cell(row=2, column=c)
        if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
            cell.fill = fill_jaune
        cell.border = border_thin
    ws.row_dimensions[2].height = 18

    # ── LIGNES DE DONNÉES à partir de la ligne 3 ──
    lignes = fiche_data.get("lignes_fiche", [])
    prix_kg = fiche_data.get("prix_kg", 0)

    for i, ligne in enumerate(lignes):
        row_num = i + 3
        # Alternance orange/blanc (identique VBA)
        fill_row = fill_orange if row_num % 2 == 1 else fill_blanc

        # Date livraison
        dv = None
        date_str = ligne.get("date", "")
        if date_str:
            try:
                dv = datetime.strptime(str(date_str), "%Y-%m-%d").date() \
                     if "-" in str(date_str) else datetime.strptime(str(date_str), "%d/%m/%Y").date()
            except:
                dv = date_str

        poids  = ligne.get("poids", 0)
        nb_sacs = ligne.get("nb_sacs", 0) or \
                  int((poids + 74) / 75) if poids else 0  # RoundUp(poids/75)
        montant = poids * prix_kg if prix_kg else ligne.get("montant", 0)

        vals = [
            ligne.get("code", ""),
            ligne.get("parc", ligne.get("parcelle", "")),
            ligne.get("nom", ""),
            ligne.get("sect", ligne.get("section", "")),
            poids,
            nb_sacs,
            fiche_data.get("fiche_id", ""),
            dv,
            int(montant) if montant else "",
            ligne.get("lon", ligne.get("longitude", "")),
            ligne.get("lat", ligne.get("latitude", ""))
        ]

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill      = fill_row
            cell.font      = font_normal
            cell.border    = border_thin
            # Format date
            if col == 8 and dv and hasattr(dv, 'strftime'):
                cell.number_format = "DD/MM/YYYY"
                cell.alignment = align_center
            elif col in (5, 6, 9):
                cell.alignment = align_center
            else:
                cell.alignment = align_left
        ws.row_dimensions[row_num].height = 16

    # ── Largeurs de colonnes ──
    widths = {1:16, 2:18, 3:20, 4:12, 5:11, 6:8, 7:22, 8:14, 9:16, 10:13, 11:12}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Figer la ligne 1 ──
    ws.freeze_panes = "A3"

    # ── Retourner les bytes ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generer_nom_fichier(fiche_data: dict) -> str:
    fiche_id = fiche_data.get("fiche_id", "fiche")
    periode  = fiche_data.get("periode", "").replace(" ", "_")
    today    = date.today().strftime("%Y%m%d")
    return f"{fiche_id}_{periode}_{today}.xlsx"
