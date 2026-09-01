"""
FichePro Manager - Moteur métier Python
Réécriture complète de la logique VBA en Python
"""
import random
import math
from datetime import date, timedelta, datetime
from typing import List, Dict, Optional, Tuple
import openpyxl
from openpyxl import load_workbook

# ============================================================
#  CONSTANTES METIER (identiques au VBA)
# ============================================================
PCT_GT        = 0.8
PCT_PT        = 0.2
LIMITE_FICHE  = 0.30
MAX_PROD      = 1000
POIDS_MIN     = 25
INTERV_JOURS  = 14

# Colonnes Base Parcelles (index 0-based)
COL_BP_CODE   = 0
COL_BP_PARC   = 1
COL_BP_NOM    = 2
COL_BP_SECT   = 3
COL_BP_ESTIM  = 8
COL_BP_STK_GT = 10
COL_BP_STK_PT = 11
COL_BP_LIV_GT = 12
COL_BP_LIV_PT = 13
COL_BP_LON    = 14
COL_BP_LAT    = 15

# Colonnes Suivi Livraisons (index 0-based)
COL_SL_CODE   = 0
COL_SL_POIDS  = 1
COL_SL_DATE   = 2
COL_SL_ID     = 3
COL_SL_PARC   = 4
COL_SL_SECT   = 5
COL_SL_PERIOD = 6
COL_SL_PRIX   = 7
COL_SL_MONT   = 8

# ============================================================
#  STRUCTURE DE DONNEES
# ============================================================
class ParcelleInfo:
    def __init__(self):
        self.parcelle   = ""
        self.producteur = ""  # code
        self.nom        = ""
        self.section    = ""
        self.lig        = 0   # index 0-based dans la liste
        self.taux       = 0.0
        self.score      = 0.0
        self.stk        = 0
        self.lon        = None
        self.lat        = None

class FicheEngine:
    """Moteur de génération de fiches - equivalent du module VBA"""

    def __init__(self, wb_path: str):
        self.wb_path         = wb_path
        self.wb              = None
        self.base_parcelles  = []   # liste de dict
        self.suivi_livraisons= []
        self.prix_kg         = 0.0
        self.calendrier      = ""   # "NOUVEAU" ou "ANCIEN"
        self.limite_actuelle = LIMITE_FICHE
        self.dict_lignes     = {}   # code -> index dans base_parcelles
        self.dict_last_liv   = {}   # parcelle -> derniere date livraison
        self.dict_dernier_sel= {}   # code -> derniere date selection
        self.densite_hebdo   = [0] * 54  # index 1-53
        self.cpt_fiche       = 0
        self.log             = []   # journal en temps reel

    def _log(self, msg: str, level: str = "info"):
        entry = {"time": datetime.now().strftime("%H:%M:%S"), "msg": msg, "level": level}
        self.log.append(entry)

    def charger_workbook(self):
        """Ouvre le fichier Excel FichePro"""
        try:
            self.wb = load_workbook(self.wb_path, keep_vba=True, data_only=True)
            self._log("Fichier Excel chargé avec succès")
            return True
        except Exception as e:
            self._log(f"Erreur chargement fichier : {e}", "error")
            return False

    def _lire_feuille(self, nom: str) -> Optional[object]:
        if self.wb is None:
            return None
        if nom not in self.wb.sheetnames:
            return None
        return self.wb[nom]

    def _val(self, cell) -> any:
        """Lecture sécurisée d'une cellule"""
        if cell is None:
            return None
        v = cell.value
        return v if v is not None else None

    def _long(self, v, default=0) -> int:
        try:
            return int(float(str(v))) if v not in (None, "") else default
        except:
            return default

    def _float(self, v, default=0.0) -> float:
        try:
            return float(str(v)) if v not in (None, "") else default
        except:
            return default

    # ============================================================
    #  INITIALISATION
    # ============================================================
    def init_dict_lignes(self):
        ws = self._lire_feuille("Base Parcelles")
        if ws is None:
            return
        self.dict_lignes = {}
        self.base_parcelles = []
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for i, row in enumerate(rows):
            if not row or row[COL_BP_CODE] in (None, ""):
                continue
            code = str(row[COL_BP_CODE])
            self.dict_lignes[code] = i
            self.base_parcelles.append({
                "code"  : code,
                "parc"  : str(row[COL_BP_PARC]) if row[COL_BP_PARC] else "",
                "nom"   : str(row[COL_BP_NOM])  if row[COL_BP_NOM]  else "",
                "sect"  : str(row[COL_BP_SECT]) if row[COL_BP_SECT] else "",
                "estim" : self._long(row[COL_BP_ESTIM]),
                "stk_gt": self._long(row[COL_BP_STK_GT]),
                "stk_pt": self._long(row[COL_BP_STK_PT]),
                "liv_gt": self._long(row[COL_BP_LIV_GT]),
                "liv_pt": self._long(row[COL_BP_LIV_PT]),
                "lon"   : self._float(row[COL_BP_LON]),
                "lat"   : self._float(row[COL_BP_LAT]),
            })
        self._log(f"{len(self.base_parcelles)} producteurs chargés")

    def precharger_memoire(self):
        ws = self._lire_feuille("Suivi Livraisons")
        if ws is None:
            return
        self.dict_last_liv    = {}
        self.dict_dernier_sel = {}
        self.densite_hebdo    = [0] * 54
        d_min = date.today() - timedelta(days=365)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row:
                continue
            parc = str(row[COL_SL_PARC]) if row[COL_SL_PARC] else ""
            cp   = str(row[COL_SL_CODE]) if row[COL_SL_CODE] else ""
            if not parc or not cp:
                continue
            rv = row[COL_SL_DATE]
            if rv is None:
                continue
            try:
                if isinstance(rv, (date, datetime)):
                    dl = rv.date() if isinstance(rv, datetime) else rv
                else:
                    dl = pd.to_datetime(rv).date()
            except:
                continue
            if parc not in self.dict_last_liv or dl > self.dict_last_liv[parc]:
                self.dict_last_liv[parc] = dl
            if cp not in self.dict_dernier_sel or dl > self.dict_dernier_sel[cp]:
                self.dict_dernier_sel[cp] = dl
            if dl >= d_min:
                sem = self._num_semaine(d_min, dl)
                if 1 <= sem <= 53:
                    self.densite_hebdo[sem] += 1
        self._log("Mémoire historique chargée")

    def init_compteur_fiches(self):
        ws = self._lire_feuille("Suivi Livraisons")
        if ws is None:
            return
        mx = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[COL_SL_ID] is None:
                continue
            s = str(row[COL_SL_ID])
            if s.startswith("FA") and len(s) >= 13:
                try:
                    n = int(s[11:15])
                    if n > mx:
                        mx = n
                except:
                    pass
        self.cpt_fiche = mx

    def nouvel_id_fiche(self) -> str:
        self.cpt_fiche += 1
        return f"FA{date.today().strftime('%Y%m%d')}-{self.cpt_fiche:04d}"

    # ============================================================
    #  PERIODE DE TRAITE
    # ============================================================
    def periode_traite(self, d_ref: date) -> Tuple[str, date, date]:
        m, a = d_ref.month, d_ref.year
        if self.calendrier == "NOUVEAU":
            if m >= 9 or m <= 2:
                if m >= 9:
                    d_deb = date(a, 9, 1)
                    d_fin = date(a+1, 2, 28) if not self._is_leap(a+1) else date(a+1, 2, 29)
                    return f"Grande traite {a}-{a+1}", d_deb, d_fin
                else:
                    d_deb = date(a-1, 9, 1)
                    d_fin = date(a, 2, 28) if not self._is_leap(a) else date(a, 2, 29)
                    return f"Grande traite {a-1}-{a}", d_deb, d_fin
            else:
                return f"Petite traite {a}", date(a, 3, 1), date(a, 8, 31)
        else:  # ANCIEN
            if m >= 10 or m <= 4:
                if m >= 10:
                    return f"Grande traite {a}-{a+1}", date(a, 10, 1), date(a+1, 4, 30)
                else:
                    return f"Grande traite {a-1}-{a}", date(a-1, 10, 1), date(a, 4, 30)
            else:
                return f"Petite traite {a}", date(a, 5, 1), date(a, 9, 30)

    def _is_leap(self, y: int) -> bool:
        return y % 4 == 0 and (y % 100 != 0 or y % 400 == 0)

    def est_grande_traite(self, periode: str) -> bool:
        return "Grande traite" in periode

    # ============================================================
    #  STOCK
    # ============================================================
    def stock_prod(self, bp: dict, periode: str) -> int:
        if self.est_grande_traite(periode):
            return bp["stk_gt"]
        return bp["stk_gt"] + bp["stk_pt"]

    def stock_total(self, periode: str, sections: List[str] = None) -> int:
        total = 0
        for bp in self.base_parcelles:
            if bp["estim"] <= 0:
                continue
            if sections and bp["sect"] not in sections:
                continue
            total += self.stock_prod(bp, periode)
        return total

    def maj_stock(self, bp: dict, poids: int, periode: str):
        if self.est_grande_traite(periode):
            bp["stk_gt"] = max(0, bp["stk_gt"] - poids)
            bp["liv_gt"] += poids
        else:
            spt = bp["stk_pt"]
            if poids <= spt:
                bp["stk_pt"] -= poids
            else:
                bp["stk_pt"] = 0
                bp["stk_gt"] = max(0, bp["stk_gt"] - (poids - spt))
            bp["liv_pt"] += poids

    # ============================================================
    #  ACTUALISER STOCKS depuis Suivi Livraisons
    # ============================================================
    def actualiser_stocks(self):
        ws = self._lire_feuille("Suivi Livraisons")
        if ws is None:
            return
        # Remettre à zéro les livraisons
        for bp in self.base_parcelles:
            bp["liv_gt"] = 0
            bp["liv_pt"] = 0
        # Recumuler depuis Suivi Livraisons
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            cp  = str(row[COL_SL_CODE])  if row[COL_SL_CODE]  else ""
            per = str(row[COL_SL_PERIOD]) if row[COL_SL_PERIOD] else ""
            pds = self._long(row[COL_SL_POIDS])
            if not cp or pds <= 0:
                continue
            if cp in self.dict_lignes:
                bp = self.base_parcelles[self.dict_lignes[cp]]
                if self.est_grande_traite(per):
                    bp["liv_gt"] += pds
                else:
                    bp["liv_pt"] += pds
        # Recalculer les stocks
        for bp in self.base_parcelles:
            est = bp["estim"]
            if est > 0:
                bp["stk_gt"] = max(0, int(est * PCT_GT) - bp["liv_gt"])
                bp["stk_pt"] = max(0, int(est * PCT_PT) - bp["liv_pt"])
        self._log(f"Stocks actualisés — {len(self.base_parcelles)} producteurs")

    # ============================================================
    #  VERIFICATION VOLUME
    # ============================================================
    def verifier_volume_disponible(self, bp: dict, poids: int, periode: str,
                                    dict_fiche: dict) -> bool:
        stk = self.stock_prod(bp, periode)
        if poids > stk:
            return False
        est = bp["estim"]
        if self.est_grande_traite(periode):
            deja = bp["liv_gt"]
            quota_max = int(est * PCT_GT)
        else:
            deja = bp["liv_gt"] + bp["liv_pt"]
            quota_max = est
        restant = quota_max - deja
        if restant > 0 and poids > restant:
            return False
        return True

    # ============================================================
    #  SCORE PRIORITE
    # ============================================================
    def calculer_score(self, bp: dict, periode: str) -> float:
        est = bp["estim"]
        stk = self.stock_prod(bp, periode)
        code = bp["code"]
        if self.est_grande_traite(periode):
            livre = bp["liv_gt"]
        else:
            livre = bp["liv_gt"] + bp["liv_pt"]
        taux = livre / est if est > 0 else 1.0
        # Composante 1 : taux non utilisé (50%)
        taux_score = (1 - taux) * 0.5
        # Composante 2 : volume restant absolu (30%)
        volume_score = min(1.0, stk / MAX_PROD) * 0.3
        # Composante 3 : disponibilité temporelle (20%)
        jours_ecart = INTERV_JOURS * 12  # bonus max si jamais sélectionné
        if code in self.dict_dernier_sel:
            delta = (date.today() - self.dict_dernier_sel[code]).days
            jours_ecart = max(0, delta)
        score_disp = min(1.0, jours_ecart / (INTERV_JOURS * 3)) * 0.2
        return taux_score + volume_score + score_disp

    # ============================================================
    #  PLAFOND PRODUCTEUR
    # ============================================================
    def plafond_producteur(self, bp: dict, poids_restant: int,
                            periode: str, dict_fiche: dict) -> int:
        stk = self.stock_prod(bp, periode)
        if stk < 1:
            return 0
        est = bp["estim"]
        if est < 1:
            return 0
        if self.est_grande_traite(periode):
            quota = int(est * PCT_GT) - bp["liv_gt"]
        else:
            quota = est - bp["liv_gt"] - bp["liv_pt"]
        if quota < 1:
            return 0
        code = bp["code"]
        deja_prod = dict_fiche.get(code, 0)
        lim30  = int(stk * self.limite_actuelle)
        libre30 = lim30 - deja_prod
        if libre30 < 1:
            return 0
        plafond = min(stk, quota, libre30, poids_restant, MAX_PROD)
        return max(0, plafond)

    # ============================================================
    #  DATE LIVRAISON OPTIMALE — tirage aléatoire
    # ============================================================
    def _num_semaine(self, d_min: date, d: date) -> int:
        return (d - d_min).days // 7 + 1

    def date_liv_optimale(self, d_debut: date, d_ref: date,
                           parcelle: str, dict_parc_dates: dict) -> date:
        d_min_hist = date.today() - timedelta(days=365)
        derniere = None
        if parcelle in self.dict_last_liv:
            derniere = self.dict_last_liv[parcelle]
        if parcelle in dict_parc_dates:
            dp = dict_parc_dates[parcelle]
            if derniere is None or dp > derniere:
                derniere = dp
        contrainte = d_debut
        if derniere:
            ideal = derniere + timedelta(days=INTERV_JOURS)
            if ideal > contrainte:
                contrainte = ideal
        if contrainte > d_ref:
            return d_ref
        # Calculer charge minimale
        min_charge = 999999
        d = contrainte
        while d <= d_ref:
            sem = self._num_semaine(d_min_hist, d)
            charge = self.densite_hebdo[sem] if 1 <= sem <= 53 else 0
            if charge < min_charge:
                min_charge = charge
            d += timedelta(days=1)
        # Collecter tous les candidats de charge minimale
        candidats = []
        d = contrainte
        while d <= d_ref:
            sem = self._num_semaine(d_min_hist, d)
            charge = self.densite_hebdo[sem] if 1 <= sem <= 53 else 0
            if charge == min_charge:
                candidats.append(d)
            d += timedelta(days=1)
        if not candidats:
            return contrainte
        return random.choice(candidats)

    def intervall_ok(self, parcelle: str, d_prop: date, dict_parc_dates: dict) -> bool:
        derniere = None
        if parcelle in self.dict_last_liv:
            derniere = self.dict_last_liv[parcelle]
        if parcelle in dict_parc_dates:
            dp = dict_parc_dates[parcelle]
            if derniere is None or dp > derniere:
                derniere = dp
        if derniere is None:
            return True
        return (d_prop - derniere).days >= INTERV_JOURS

    # ============================================================
    #  CALCUL PALIER MINIMUM
    # ============================================================
    def calculer_palier_minimum(self, periode: str, poids_restant: int,
                                 parcelles_actives: List[dict],
                                 dict_fiche: dict) -> float:
        limite_pct = int(self.limite_actuelle * 100)
        for palier_pct in range(limite_pct + 1, 101):
            palier = palier_pct / 100.0
            vol_total = 0
            for bp in parcelles_actives:
                stk = self.stock_prod(bp, periode)
                if stk < 1:
                    continue
                est = bp["estim"]
                if est < 1:
                    continue
                if self.est_grande_traite(periode):
                    quota = int(est * PCT_GT) - bp["liv_gt"]
                else:
                    quota = est - bp["liv_gt"] - bp["liv_pt"]
                if quota < 1:
                    continue
                code = bp["code"]
                deja = dict_fiche.get(code, 0)
                lim  = int(stk * palier)
                libre = lim - deja
                if libre < 1:
                    continue
                vol_total += min(stk, quota, libre, MAX_PROD)
            if vol_total >= poids_restant:
                return palier
        return -1.0

    # ============================================================
    #  TRANSFERT GT -> PT
    # ============================================================
    def transfert_gt_pt(self):
        for bp in self.base_parcelles:
            if bp["stk_gt"] > 0:
                bp["stk_pt"] += bp["stk_gt"]
                bp["stk_gt"] = 0
        self._log("Transfert GT → PT effectué")

    # ============================================================
    #  GENERATION DE FICHE — cœur du moteur
    # ============================================================
    def generer_fiche(self, poids_total: int, d_ref: date,
                       d_debut: date, sections: List[str],
                       prix_kg: float) -> dict:
        """
        Génère une fiche complète.
        Retourne un dict avec les lignes, l'ID, les stats, les warnings.
        """
        self.limite_actuelle = LIMITE_FICHE
        self.prix_kg = prix_kg
        periode, d_deb_camp, _ = self.periode_traite(d_ref)
        fiche_id = self.nouvel_id_fiche()
        self._log(f"Génération fiche {fiche_id} — {periode}")

        # Filtrer les parcelles éligibles dans les sections choisies
        eligibles = [
            bp for bp in self.base_parcelles
            if bp["sect"] in sections
            and bp["estim"] > 0
            and self.stock_prod(bp, periode) >= POIDS_MIN
        ]
        if not eligibles:
            return {"erreur": "Aucune parcelle avec stock suffisant dans ces sections."}

        # Trier par score décroissant
        for bp in eligibles:
            bp["_score"] = self.calculer_score(bp, periode)
        import random; random.shuffle(eligibles)
        eligibles.sort(key=lambda x: x["_score"], reverse=True)

        # Vérifier volume sections
        vol_sect = sum(self.stock_prod(bp, periode) for bp in eligibles)
        if vol_sect < poids_total:
            return {
                "erreur": f"Volume insuffisant dans les sections ({vol_sect:,} kg disponibles, {poids_total:,} kg demandés)."
            }

        # Snapshot pour rollback
        snapshot = {
            bp["code"]: {
                "stk_gt": bp["stk_gt"], "stk_pt": bp["stk_pt"],
                "liv_gt": bp["liv_gt"], "liv_pt": bp["liv_pt"]
            }
            for bp in eligibles
        }

        dict_fiche       = {}
        dict_parc_dates  = {}
        lignes_fiche     = []
        lignes_suivi     = []
        poids_restant    = poids_total
        tour             = 0
        auto_saut_intv   = False
        warnings         = []

        while poids_restant > 0 and tour < 300:
            avant_tour    = poids_restant
            tour         += 1
            bloc_lim30    = False
            bloc_intv     = False

            for bp in eligibles:
                if poids_restant < POIDS_MIN:
                    break

                code  = bp["code"]
                parc  = bp["parc"]

                # Vérif limite fiche
                deja   = dict_fiche.get(code, 0)
                lim30c = int(self.stock_prod(bp, periode) * self.limite_actuelle)
                if deja >= lim30c:
                    bloc_lim30 = True
                    continue

                # Plafond
                plaf = self.plafond_producteur(bp, poids_restant, periode, dict_fiche)
                if plaf < 1:
                    continue

                # Poids = maximum possible
                p_choisi = min(plaf, poids_restant)
                if p_choisi < POIDS_MIN:
                    if p_choisi > 0:
                        self._absorber_residu(
                            p_choisi, fiche_id, periode, eligibles,
                            dict_fiche, lignes_fiche, lignes_suivi
                        )
                        poids_restant -= p_choisi
                    continue

                # Date livraison
                d_parc = self.date_liv_optimale(d_debut, d_ref, parc, dict_parc_dates)
                if not self.intervall_ok(parc, d_parc, dict_parc_dates):
                    if not auto_saut_intv:
                        bloc_intv = True
                        continue

                # Enregistrer
                self._enregistrer_ligne(
                    bp, parc, p_choisi, fiche_id, periode, d_parc,
                    dict_fiche, lignes_fiche, lignes_suivi, dict_parc_dates
                )
                poids_restant -= p_choisi

            # Residu < POIDS_MIN
            if 0 < poids_restant < POIDS_MIN:
                absorbe = self._absorber_residu(
                    poids_restant, fiche_id, periode, eligibles,
                    dict_fiche, lignes_fiche, lignes_suivi
                )
                if absorbe:
                    poids_restant = 0

            # Blocage total
            if poids_restant > 0 and poids_restant == avant_tour:
                if bloc_lim30:
                    # Calculer palier minimum
                    palier_min = self.calculer_palier_minimum(
                        periode, poids_restant, eligibles, dict_fiche
                    )
                    if palier_min < 0:
                        # Rollback
                        self._rollback(snapshot, eligibles)
                        return {
                            "erreur": f"Volume restant de {poids_restant:,} kg non distribuable même en levant la limite."
                        }
                    # Retourner une demande d'autorisation à l'interface
                    return {
                        "autorisation_requise": "limite_fiche",
                        "limite_actuelle"     : self.limite_actuelle,
                        "palier_min"          : palier_min,
                        "poids_restant"       : poids_restant,
                        "etat_partiel"        : {
                            "lignes_fiche"   : lignes_fiche,
                            "lignes_suivi"   : lignes_suivi,
                            "dict_fiche"     : dict_fiche,
                            "dict_parc_dates": dict_parc_dates,
                            "poids_restant"  : poids_restant,
                            "tour"           : tour,
                            "snapshot"       : snapshot,
                            "eligibles_codes": [bp["code"] for bp in eligibles]
                        }
                    }
                elif bloc_intv:
                    auto_saut_intv = True
                    if auto_saut_intv:
                        warnings.append("Intervalle minimum raccourci (autorisé)")
                else:
                    break

        # Vérification intégrité
        total_v = sum(l["poids"] for l in lignes_fiche)
        if total_v != poids_total:
            self._rollback(snapshot, eligibles)
            return {"erreur": f"Écart non résolu : {poids_total - total_v:,} kg manquants."}

        if auto_saut_intv:
            warnings.append("Intervalle entre livraisons raccourci")

        return {
            "succes"      : True,
            "fiche_id"    : fiche_id,
            "periode"     : periode,
            "prix_kg"     : prix_kg,
            "poids_total" : poids_total,
            "nb_producteurs": len(set(l["code"] for l in lignes_fiche)),
            "nb_sacs"     : sum(math.ceil(l["poids"] / 75) for l in lignes_fiche),
            "montant"     : int(poids_total * prix_kg),
            "lignes_fiche": lignes_fiche,
            "lignes_suivi": lignes_suivi,
            "warnings"    : warnings,
            "d_debut"     : d_debut.isoformat(),
            "d_ref"       : d_ref.isoformat()
        }

    def continuer_apres_autorisation(self, palier_accepte: float, etat: dict,
                                      periode: str, d_debut: date, d_ref: date,
                                      prix_kg: float, poids_total: int) -> dict:
        """Reprend la génération après autorisation de l'utilisateur"""
        self.limite_actuelle = palier_accepte
        # Restaurer l'état
        codes = etat["eligibles_codes"]
        eligibles = [bp for bp in self.base_parcelles if bp["code"] in codes]
        eligibles.sort(key=lambda x: x.get("_score", 0), reverse=True)
        # Relancer avec le nouvel état
        # (simplification : relancer complètement avec la nouvelle limite)
        dict_fiche       = etat["dict_fiche"]
        dict_parc_dates  = etat["dict_parc_dates"]
        lignes_fiche     = etat["lignes_fiche"]
        lignes_suivi     = etat["lignes_suivi"]
        poids_restant    = etat["poids_restant"]
        tour             = etat["tour"]
        fiche_id         = lignes_fiche[0]["id"] if lignes_fiche else self.nouvel_id_fiche()

        # Continuer la boucle
        while poids_restant > 0 and tour < 300:
            avant_tour = poids_restant
            tour += 1
            for bp in eligibles:
                if poids_restant < POIDS_MIN:
                    break
                code  = bp["code"]
                parc  = bp["parc"]
                deja  = dict_fiche.get(code, 0)
                lim30c = int(self.stock_prod(bp, periode) * self.limite_actuelle)
                if deja >= lim30c:
                    continue
                plaf = self.plafond_producteur(bp, poids_restant, periode, dict_fiche)
                if plaf < 1:
                    continue
                p_choisi = min(plaf, poids_restant)
                if p_choisi < POIDS_MIN:
                    continue
                d_parc = self.date_liv_optimale(d_debut, d_ref, parc, dict_parc_dates)
                self._enregistrer_ligne(
                    bp, parc, p_choisi, fiche_id, periode, d_parc,
                    dict_fiche, lignes_fiche, lignes_suivi, dict_parc_dates
                )
                poids_restant -= p_choisi
            if 0 < poids_restant < POIDS_MIN:
                self._absorber_residu(
                    poids_restant, fiche_id, periode, eligibles,
                    dict_fiche, lignes_fiche, lignes_suivi
                )
                poids_restant = 0
            if poids_restant == avant_tour:
                break

        total_v = sum(l["poids"] for l in lignes_fiche)
        return {
            "succes"        : True,
            "fiche_id"      : fiche_id,
            "periode"       : periode,
            "prix_kg"       : prix_kg,
            "poids_total"   : poids_total,
            "nb_producteurs": len(set(l["code"] for l in lignes_fiche)),
            "nb_sacs"       : sum(math.ceil(l["poids"] / 75) for l in lignes_fiche),
            "montant"       : int(total_v * prix_kg),
            "lignes_fiche"  : lignes_fiche,
            "lignes_suivi"  : lignes_suivi,
            "warnings"      : [f"Limite fiche portée à {palier_accepte:.0%} (autorisée)"],
            "d_debut"       : d_debut.isoformat(),
            "d_ref"         : d_ref.isoformat()
        }

    def _enregistrer_ligne(self, bp, parc, poids, fiche_id, periode, d_parc,
                            dict_fiche, lignes_fiche, lignes_suivi, dict_parc_dates):
        code    = bp["code"]
        montant = int(poids * self.prix_kg)
        nb_sacs = math.ceil(poids / 75)
        lignes_fiche.append({
            "code"    : code,
            "parc"    : parc,
            "nom"     : bp["nom"],
            "sect"    : bp["sect"],
            "poids"   : poids,
            "nb_sacs" : nb_sacs,
            "id"      : fiche_id,
            "date"    : d_parc.isoformat(),
            "prix"    : self.prix_kg,
            "montant" : montant,
            "lon"     : bp["lon"],
            "lat"     : bp["lat"]
        })
        lignes_suivi.append({
            "code"   : code, "poids"  : poids,
            "date"   : d_parc.isoformat(), "id": fiche_id,
            "parc"   : parc, "sect"   : bp["sect"],
            "periode": periode, "prix" : self.prix_kg,
            "montant": montant
        })
        self.maj_stock(bp, poids, periode)
        dict_fiche[code]    = dict_fiche.get(code, 0) + poids
        dict_parc_dates[parc] = d_parc
        sem = self._num_semaine(date.today() - timedelta(days=365), d_parc)
        if 1 <= sem <= 53:
            self.densite_hebdo[sem] += 1

    def _absorber_residu(self, residu, fiche_id, periode, eligibles,
                          dict_fiche, lignes_fiche, lignes_suivi) -> bool:
        # T1 : producteur non encore dans la fiche
        candidats = [
            bp for bp in eligibles
            if bp["code"] not in dict_fiche
            and self.verifier_volume_disponible(bp, residu, periode, dict_fiche)
        ]
        import random; random.shuffle(candidats)
        candidats.sort(key=lambda x: x.get("_score", 0), reverse=True)
        for bp in candidats:
            d_parc = date.today()
            self._enregistrer_ligne(
                bp, bp["parc"], residu, fiche_id, periode, d_parc,
                dict_fiche, lignes_fiche, lignes_suivi, {}
            )
            return True
        # T2 : augmenter un existant
        for bp in eligibles:
            code = bp["code"]
            if code not in dict_fiche:
                continue
            plaf = self.plafond_producteur(bp, residu, periode, dict_fiche)
            if plaf < residu:
                continue
            if not self.verifier_volume_disponible(bp, residu, periode, dict_fiche):
                continue
            # Trouver la ligne dans lignes_fiche
            for lg in lignes_fiche:
                if lg["code"] == code and lg["id"] == fiche_id:
                    lg["poids"]   += residu
                    lg["montant"] += int(residu * self.prix_kg)
                    lg["nb_sacs"]  = math.ceil(lg["poids"] / 75)
                    self.maj_stock(bp, residu, periode)
                    dict_fiche[code] += residu
                    return True
        return False

    def _rollback(self, snapshot: dict, eligibles: List[dict]):
        for bp in eligibles:
            code = bp["code"]
            if code in snapshot:
                bp["stk_gt"] = snapshot[code]["stk_gt"]
                bp["stk_pt"] = snapshot[code]["stk_pt"]
                bp["liv_gt"] = snapshot[code]["liv_gt"]
                bp["liv_pt"] = snapshot[code]["liv_pt"]

    # ============================================================
    #  PERSISTANCE — sauvegarder dans le fichier Excel
    # ============================================================
    def sauvegarder_fiche(self, result: dict) -> bool:
        """Ecrit la fiche générée dans le fichier Excel"""
        try:
            wb = load_workbook(self.wb_path, keep_vba=True)
            # --- Suivi Livraisons ---
            ws_sl = wb["Suivi Livraisons"]
            lr = ws_sl.max_row
            for lg in result["lignes_suivi"]:
                lr += 1
                ws_sl.cell(lr, 1, lg["code"])
                ws_sl.cell(lr, 2, lg["poids"])
                ws_sl.cell(lr, 3, lg["date"])
                ws_sl.cell(lr, 4, lg["id"])
                ws_sl.cell(lr, 5, lg["parc"])
                ws_sl.cell(lr, 6, lg["sect"])
                ws_sl.cell(lr, 7, lg["periode"])
                ws_sl.cell(lr, 8, lg["prix"])
                ws_sl.cell(lr, 9, lg["montant"])
            # --- Base Parcelles stocks ---
            ws_bp = wb["Base Parcelles"]
            for row in ws_bp.iter_rows(min_row=2):
                code = row[COL_BP_CODE].value
                if code and code in self.dict_lignes:
                    bp = self.base_parcelles[self.dict_lignes[code]]
                    row[COL_BP_STK_GT].value = bp["stk_gt"]
                    row[COL_BP_STK_PT].value = bp["stk_pt"]
                    row[COL_BP_LIV_GT].value = bp["liv_gt"]
                    row[COL_BP_LIV_PT].value = bp["liv_pt"]
            wb.save(self.wb_path)
            self._log(f"Fiche {result['fiche_id']} sauvegardée dans Excel")
            return True
        except Exception as e:
            self._log(f"Erreur sauvegarde : {e}", "error")
            return False

    # ============================================================
    #  IMPORT REGISTRE RAINFOREST ALLIANCE
    # ============================================================
    def importer_registre_ra(self, ra_path: str) -> dict:
        """
        Lit le registre RA (Annexe S13 Rainforest Alliance).
        
        Logique d'estimation par parcelle :
          Estimation parcelle = Rendement estimé/ha (onglet 2) × Superficie parcelle (onglet 3)
          STK_GT = estimation × 80%
          STK_PT = estimation × 20%
        
        Un producteur peut avoir plusieurs parcelles — chacune a sa propre estimation.
        
        Structure vérifiée sur fr_GMR_template_2026-2027 :
        Onglet 1 : Col0=ID | Col4=Section | Col9=Prénom | Col10=Nom
        Onglet 2 : Col0=ID | Col9=Rendement estimé/ha
        Onglet 3 : Col0=ID exploitation | Col1=ID parcelle | Col2=Superficie | Col3=Lat | Col4=Lon
        """
        try:
            try:
                wb_ra = load_workbook(ra_path, data_only=True)
            except Exception:
                wb_ra = load_workbook(ra_path, data_only=True, read_only=True)

            def find_sheet(wb, keyword):
                for name in wb.sheetnames:
                    if keyword.lower() in name.lower():
                        return wb[name]
                return None

            ws1 = find_sheet(wb_ra, "exploitation")
            ws2 = find_sheet(wb_ra, "certifi")
            ws3 = find_sheet(wb_ra, "unit")

            if not ws1:
                return {"erreur": "Onglet 'exploitation agricole' introuvable. "
                        "Vérifiez que vous importez le registre Rainforest Alliance (Annexe S13)."}
            if not ws3:
                return {"erreur": "Onglet 'Unité agricole' introuvable dans le registre RA."}

            # ── ONGLET 1 : Infos producteurs ──
            prod_data = {}
            for row in ws1.iter_rows(min_row=3, values_only=True):
                if not row or row[0] in (None, ""): continue
                iid = str(row[0]).strip()
                if any(kw in iid.lower() for kw in ["identifiant", "exploitation", "interne"]):
                    continue
                prenom      = str(row[9]).strip()  if row[9]  else ""
                nom_famille = str(row[10]).strip() if row[10] else ""
                nom_complet = f"{prenom} {nom_famille}".strip() or iid
                section     = str(row[4]).strip()  if row[4]  else ""
                prod_data[iid] = {
                    "nom"          : nom_complet,
                    "sect"         : section,
                    "rendement_ha" : 0.0,
                    "parcelles"    : []
                }

            if not prod_data:
                return {"erreur": "Aucun producteur trouvé dans l'onglet 'exploitation agricole'."}

            # ── ONGLET 2 : Rendement estimé/ha par exploitation ──
            if ws2:
                for row in ws2.iter_rows(min_row=3, values_only=True):
                    if not row or row[0] in (None, ""): continue
                    iid = str(row[0]).strip()
                    if iid not in prod_data: continue
                    try:
                        rha = float(row[9]) if row[9] not in (None, "") else 0.0
                    except:
                        rha = 0.0
                    prod_data[iid]["rendement_ha"] = rha

            # ── ONGLET 3 : Parcelles avec superficie et GPS ──
            # Estimation parcelle = rendement_ha × superficie_parcelle
            for row in ws3.iter_rows(min_row=3, values_only=True):
                if not row or row[0] in (None, ""): continue
                iid     = str(row[0]).strip()
                parc_id = str(row[1]).strip() if row[1] else f"{iid}-P1"
                if iid not in prod_data: continue
                try:
                    superficie = float(row[2]) if row[2] not in (None, "") else 0.0
                except:
                    superficie = 0.0
                try:
                    lat = float(row[3]) if row[3] not in (None, "") else None
                except:
                    lat = None
                try:
                    lon = float(row[4]) if row[4] not in (None, "") else None
                except:
                    lon = None
                if superficie <= 0: continue
                rha   = prod_data[iid]["rendement_ha"]
                estim = int(rha * superficie)
                if estim <= 0: continue
                prod_data[iid]["parcelles"].append({
                    "parc_id"   : parc_id,
                    "estim"     : estim,
                    "stk_gt"    : int(estim * PCT_GT),
                    "stk_pt"    : int(estim * PCT_PT),
                    "lat"       : lat,
                    "lon"       : lon
                })

            # Charger directement en mémoire — pas besoin de fichier Excel
            parc_existantes = {bp["parc"]: i for i, bp in enumerate(self.base_parcelles)}
            ajoutes    = 0
            mis_a_jour = 0
            total_parc = 0

            for iid, p in prod_data.items():
                if not p["parcelles"]: continue
                for pa in p["parcelles"]:
                    total_parc += 1
                    pid = pa["parc_id"]
                    bp_new = {
                        "code"  : iid,  "parc"  : pid,
                        "nom"   : p["nom"], "sect": p["sect"],
                        "estim" : pa["estim"],
                        "stk_gt": pa["stk_gt"], "stk_pt": pa["stk_pt"],
                        "liv_gt": 0, "liv_pt": 0,
                        "lon"   : pa["lon"], "lat": pa["lat"]
                    }
                    if pid in parc_existantes:
                        self.base_parcelles[parc_existantes[pid]].update(bp_new)
                        mis_a_jour += 1
                    else:
                        idx = len(self.base_parcelles)
                        self.base_parcelles.append(bp_new)
                        self.dict_lignes[iid] = idx
                        parc_existantes[pid] = idx
                        ajoutes += 1

            self._log(f"Import RA : {ajoutes} ajoutés, {mis_a_jour} mis à jour "
                      f"({len(prod_data)} producteurs, {total_parc} parcelles)")
            return {
                "succes"    : True,
                "ajoutes"   : ajoutes,
                "mis_a_jour": mis_a_jour,
                "total"     : total_parc,
                "nb_prod"   : len(prod_data)
            }

        except Exception as e:
            self._log(f"Erreur import RA : {e}", "error")
            return {"erreur": f"Erreur lors de l'import : {str(e)}"}


    def stats_dashboard(self, periode: str = None) -> dict:
        if not self.base_parcelles:
            return {}
        if periode is None:
            periode, _, _ = self.periode_traite(date.today())
        vol_total  = self.stock_total(periode)
        nb_prod    = len(self.base_parcelles)
        nb_utilises = sum(
            1 for bp in self.base_parcelles
            if (bp["liv_gt"] + bp["liv_pt"]) > 0
        )
        sections = {}
        for bp in self.base_parcelles:
            s = bp["sect"]
            stk = self.stock_prod(bp, periode)
            if s not in sections:
                sections[s] = {"vol": 0, "nb": 0}
            sections[s]["vol"] += stk
            sections[s]["nb"]  += 1
        return {
            "vol_total"  : vol_total,
            "nb_prod"    : nb_prod,
            "nb_utilises": nb_utilises,
            "nb_fiches"  : self.cpt_fiche,
            "periode"    : periode,
            "sections"   : sections,
            "calendrier" : self.calendrier
        }
